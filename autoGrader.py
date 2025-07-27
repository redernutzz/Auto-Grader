import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import random
import pandas as pd
from datetime import datetime
import os

# ================== Subject-specific percentage weights ==================
subject_weights = {
    "Math": {"Written": 40, "Performance": 40, "Assessment": 20},
    "English": {"Written": 30, "Performance": 50, "Assessment": 20},
    "Science": {"Written": 35, "Performance": 45, "Assessment": 20},
    "Filipino": {"Written": 25, "Performance": 50, "Assessment": 25},
    "Araling Panlipunan": {"Written": 30, "Performance": 50, "Assessment": 20},
}

# ================== Grade Logic ==================
def calculate_component_grade(scores, perfect_scores):
    return (sum(scores) / sum(perfect_scores)) * 100 if sum(perfect_scores) > 0 else 0

def find_combination(w_perfect, p_perfect, a_perfect, w_weight, p_weight, a_weight, target_grade, tolerance=0.01, max_attempts=100000):
    for _ in range(max_attempts):
        w_scores = [random.randint(0, p) for p in w_perfect]
        p_scores = [random.randint(0, p) for p in p_perfect]
        a_scores = [random.randint(0, p) for p in a_perfect]

        w_grade = calculate_component_grade(w_scores, w_perfect)
        p_grade = calculate_component_grade(p_scores, p_perfect)
        a_grade = calculate_component_grade(a_scores, a_perfect)

        final = (w_grade * w_weight + p_grade * p_weight + a_grade * a_weight) / 100

        if abs(final - target_grade) <= tolerance:
            return {
                "Written Works": (w_scores, w_grade),
                "Performance Task": (p_scores, p_grade),
                "Quarterly Assessment": (a_scores, a_grade),
                "Final Grade": round(final, 2)
            }
    return None

def create_excel_file(results, subject, w_perfect, p_perfect, a_perfect, w_weight, p_weight, a_weight):
    """Create an Excel file with the grades in the specified format"""
    try:
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Excel File"
        )
        
        if not filename:
            return False
            
        # Create the Excel data structure
        data = []
        
        # Header row 1 - Component names and weights
        header1 = [''] + [f'WRITTEN WORKS ({w_weight}%)'] + [''] * (len(w_perfect) + 1) + \
                 [f'PERFORMANCE TASKS ({p_weight}%)'] + [''] * (len(p_perfect) + 1) + \
                 [f'QUARTERLY ASSESSMENT ({a_weight}%)', '', 'Initial']
        
        # Header row 2 - Activity numbers and totals
        header2 = [''] + [str(i+1) for i in range(len(w_perfect))] + ['Total', 'PS', 'WS'] + \
                 [str(i+1) for i in range(len(p_perfect))] + ['Total', 'PS', 'WS'] + \
                 ['1', 'PS', 'WS', 'Grade']
        
        # Perfect scores row
        perfect_row = [''] + w_perfect + [sum(w_perfect), 100.00, f'{w_weight}%'] + \
                     p_perfect + [sum(p_perfect), 100.00, f'{p_weight}%'] + \
                     [sum(a_perfect), 100.00, f'{a_weight}%', '']
        
        # Add headers and perfect scores
        data.append(header1)
        data.append(header2)
        data.append(perfect_row)
        
        # Add student data
        for i, result in enumerate(results, 1):
            if "Error" not in result:
                # Written Works data
                w_scores = result["Written Works"][0]
                w_grade = result["Written Works"][1]
                w_total = sum(w_scores)
                
                # Performance Tasks data
                p_scores = result["Performance Task"][0]
                p_grade = result["Performance Task"][1]
                p_total = sum(p_scores)
                
                # Assessment data
                a_scores = result["Quarterly Assessment"][0]
                a_grade = result["Quarterly Assessment"][1]
                a_total = sum(a_scores)
                
                # Final grade
                final_grade = result["Final Grade"]
                
                # Create student row
                student_row = [str(i)] + w_scores + [w_total, f'{w_grade:.2f}', f'{(w_grade * w_weight / 100):.2f}'] + \
                             p_scores + [p_total, f'{p_grade:.2f}', f'{(p_grade * p_weight / 100):.2f}'] + \
                             a_scores + [f'{a_grade:.2f}', f'{(a_grade * a_weight / 100):.2f}', f'{final_grade:.2f}']
                
                data.append(student_row)
            else:
                # Handle error case
                error_row = [str(i)] + ['ERROR'] * (len(header2) - 1)
                data.append(error_row)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Save to Excel with formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'{subject} Grades', index=False, header=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets[f'{subject} Grades']
            
            # Apply some basic formatting
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Header formatting
            header_font = Font(bold=True, size=10)
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            
            # Apply formatting to headers
            for row in range(1, 4):  # First 3 rows are headers
                for col in range(1, df.shape[1] + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 15)  # Max width of 15
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        messagebox.showinfo("Success", f"Excel file created successfully at:\n{filename}")
        return True
        
    except ImportError:
        messagebox.showerror("Error", "Required libraries not found. Please install:\npip install pandas openpyxl")
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Failed to create Excel file:\n{str(e)}")
        return False

# ================== Modern GUI Styling ==================
class ModernStyle:
    # Color scheme
    PRIMARY = "#1a1a1a"
    SECONDARY = "#2d2d2d"
    ACCENT = "#007acc"
    SUCCESS = "#28a745"
    WARNING = "#ffc107"
    ERROR = "#dc3545"
    BACKGROUND = "#f8f9fa"
    SURFACE = "#ffffff"
    TEXT_PRIMARY = "#212529"
    TEXT_SECONDARY = "#6c757d"
    BORDER = "#e9ecef"
    
    # Fonts
    FONT_TITLE = ("Segoe UI", 18, "bold")
    FONT_HEADER = ("Segoe UI", 14, "bold")
    FONT_MEDIUM = ("Segoe UI", 11)
    FONT_SMALL = ("Segoe UI", 10)
    FONT_BUTTON = ("Segoe UI", 10, "bold")

def create_modern_button(parent, text, command, bg_color=ModernStyle.ACCENT, width=12, height=1):
    """Create a compact modern styled button"""
    btn = tk.Button(
        parent, 
        text=text,
        command=command,
        font=ModernStyle.FONT_BUTTON,
        bg=bg_color,
        fg="white",
        bd=0,
        relief="flat",
        padx=15,
        pady=5,
        width=width,
        height=height,
        cursor="hand2"
    )
    
    # Hover effects
    def on_enter(e):
        btn['bg'] = ModernStyle.SECONDARY
    def on_leave(e):
        btn['bg'] = bg_color
    
    btn.bind("<Enter>", on_enter)
    btn.bind("<Leave>", on_leave)
    
    return btn

def create_modern_entry(parent, width=20):
    """Create a compact modern styled entry"""
    entry = tk.Entry(
        parent,
        font=ModernStyle.FONT_MEDIUM,
        bg=ModernStyle.SURFACE,
        fg=ModernStyle.TEXT_PRIMARY,
        bd=1,
        relief="solid",
        width=width,
        highlightthickness=1,
        highlightcolor=ModernStyle.ACCENT,
        highlightbackground=ModernStyle.BORDER
    )
    return entry

# ================== Main Application ==================
class GradeGeneratorApp:
    def __init__(self):
        self.root = tk.Tk()
        self.setup_window()
        self.entries = {}
        self.target_entries = []
        self.subject_var = tk.StringVar()
        self.generated_results = []
        
        # Initialize with first screen
        self.build_initial_form()
    
    def setup_window(self):
        """Setup main window with compact sizing"""
        self.root.title("Grade Generator Pro")
        self.root.geometry("1000x650")
        self.root.configure(bg=ModernStyle.BACKGROUND)
        self.root.resizable(False, False)
        
        # Center window on screen
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (1000 // 2)
        y = (self.root.winfo_screenheight() // 2) - (650 // 2)
        self.root.geometry(f"1000x650+{x}+{y}")
    
    def clear_window(self):
        """Clear all widgets from window"""
        for widget in self.root.winfo_children():
            widget.destroy()
    
    def build_initial_form(self):
        """Build the compact initial subject selection form"""
        self.clear_window()
        
        # Main container
        main_frame = tk.Frame(self.root, bg=ModernStyle.BACKGROUND)
        main_frame.pack(expand=True, fill="both", padx=30, pady=30)
        
        # Compact header
        title_label = tk.Label(
            main_frame,
            text="Grade Generator Pro",
            font=("Segoe UI", 24, "bold"),
            bg=ModernStyle.BACKGROUND,
            fg=ModernStyle.PRIMARY
        )
        title_label.pack(pady=(0, 5))
        
        subtitle_label = tk.Label(
            main_frame,
            text="Generate student grades with professional Excel export",
            font=ModernStyle.FONT_MEDIUM,
            bg=ModernStyle.BACKGROUND,
            fg=ModernStyle.TEXT_SECONDARY
        )
        subtitle_label.pack(pady=(0, 30))
        
        # Compact subject selection card
        card_frame = tk.Frame(main_frame, bg=ModernStyle.SURFACE, relief="solid", bd=1)
        card_frame.pack(pady=20, ipadx=40, ipady=30)
        
        # Subject selection
        subject_label = tk.Label(
            card_frame,
            text="Select Subject:",
            font=ModernStyle.FONT_HEADER,
            bg=ModernStyle.SURFACE,
            fg=ModernStyle.TEXT_PRIMARY
        )
        subject_label.pack(pady=(0, 10))
        
        # Modern combobox styling
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Modern.TCombobox',
                       fieldbackground=ModernStyle.SURFACE,
                       background=ModernStyle.SURFACE,
                       borderwidth=1,
                       focuscolor=ModernStyle.ACCENT)
        
        dropdown = ttk.Combobox(
            card_frame,
            textvariable=self.subject_var,
            state="readonly",
            font=ModernStyle.FONT_MEDIUM,
            style='Modern.TCombobox',
            width=25
        )
        dropdown["values"] = list(subject_weights.keys())
        dropdown.pack(pady=(0, 15))
        dropdown.bind("<<ComboboxSelected>>", lambda e: self.build_form())
        
        # Compact instructions
        instruction_label = tk.Label(
            card_frame,
            text="Choose a subject to begin",
            font=ModernStyle.FONT_SMALL,
            bg=ModernStyle.SURFACE,
            fg=ModernStyle.TEXT_SECONDARY
        )
        instruction_label.pack()
    
    def build_form(self):
        """Build compact form for entering grade parameters"""
        self.clear_window()
        
        # Main container
        main_frame = tk.Frame(self.root, bg=ModernStyle.BACKGROUND)
        main_frame.pack(expand=True, fill="both", padx=30, pady=20)
        
        # Compact header
        header_frame = tk.Frame(main_frame, bg=ModernStyle.BACKGROUND)
        header_frame.pack(fill="x", pady=(0, 15))
        
        back_btn = create_modern_button(
            header_frame, "← Back", self.build_initial_form, 
            bg_color=ModernStyle.TEXT_SECONDARY, width=8
        )
        back_btn.pack(side="left")
        
        title_label = tk.Label(
            header_frame,
            text=f"Setup Parameters - {self.subject_var.get()}",
            font=ModernStyle.FONT_TITLE,
            bg=ModernStyle.BACKGROUND,
            fg=ModernStyle.PRIMARY
        )
        title_label.pack(side="left", padx=(15, 0))
        
        # Form in a single compact card
        card_frame = tk.Frame(main_frame, bg=ModernStyle.SURFACE, relief="solid", bd=1)
        card_frame.pack(fill="both", expand=True, pady=10)
        
        # Content frame with padding
        content_frame = tk.Frame(card_frame, bg=ModernStyle.SURFACE)
        content_frame.pack(padx=25, pady=20, fill="both", expand=True)
        
        # Create 3-column layout for compactness
        left_frame = tk.Frame(content_frame, bg=ModernStyle.SURFACE)
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 15))
        
        middle_frame = tk.Frame(content_frame, bg=ModernStyle.SURFACE)
        middle_frame.pack(side="left", fill="both", expand=True, padx=(0, 15))
        
        right_frame = tk.Frame(content_frame, bg=ModernStyle.SURFACE)
        right_frame.pack(side="left", fill="both", expand=True)
        
        # Written Works section
        self.create_section(left_frame, "📝 Written Works", [
            ("Written Count", "Activities count"),
            ("Written Perfects", "Perfect scores (comma-separated)")
        ])
        
        # Performance Tasks section
        self.create_section(middle_frame, "🎯 Performance Tasks", [
            ("Performance Count", "Tasks count"),
            ("Performance Perfects", "Perfect scores (comma-separated)")
        ])
        
        # Assessment + Students section
        self.create_section(right_frame, "📊 Assessment & Students", [
            ("Assessment Count", "Assessments count"),
            ("Assessment Perfects", "Perfect scores (comma-separated)"),
            ("Number of Students", "Total students")
        ])
        
        # Next button at bottom
        button_frame = tk.Frame(content_frame, bg=ModernStyle.SURFACE)
        button_frame.pack(side="bottom", pady=(15, 0))
        
        next_btn = create_modern_button(button_frame, "Next →", self.build_targets, width=15)
        next_btn.pack()
    
    def create_section(self, parent, title, fields):
        """Create a compact section with fields"""
        # Section header
        title_label = tk.Label(
            parent,
            text=title,
            font=ModernStyle.FONT_HEADER,
            bg=ModernStyle.SURFACE,
            fg=ModernStyle.ACCENT
        )
        title_label.pack(anchor="w", pady=(0, 10))
        
        # Fields
        for field_name, description in fields:
            field_frame = tk.Frame(parent, bg=ModernStyle.SURFACE)
            field_frame.pack(fill="x", pady=6)
            
            label = tk.Label(
                field_frame,
                text=field_name + ":",
                font=ModernStyle.FONT_MEDIUM,
                bg=ModernStyle.SURFACE,
                fg=ModernStyle.TEXT_PRIMARY
            )
            label.pack(anchor="w")
            
            entry = create_modern_entry(field_frame, width=25)
            entry.pack(anchor="w", pady=(2, 0))
            
            desc_label = tk.Label(
                field_frame,
                text=description,
                font=ModernStyle.FONT_SMALL,
                bg=ModernStyle.SURFACE,
                fg=ModernStyle.TEXT_SECONDARY
            )
            desc_label.pack(anchor="w", pady=(1, 0))
            
            self.entries[field_name] = entry
    
    def build_targets(self):
        """Build compact target grades input form"""
        try:
            # Validate and store form data
            subject = self.subject_var.get()
            weights = subject_weights[subject]
            self.w_weight = weights["Written"]
            self.p_weight = weights["Performance"]
            self.a_weight = weights["Assessment"]

            w_count = int(self.entries["Written Count"].get())
            self.w_perfect = list(map(int, self.entries["Written Perfects"].get().split(',')))

            p_count = int(self.entries["Performance Count"].get())
            self.p_perfect = list(map(int, self.entries["Performance Perfects"].get().split(',')))

            a_count = int(self.entries["Assessment Count"].get())
            self.a_perfect = list(map(int, self.entries["Assessment Perfects"].get().split(',')))

            self.num_students = int(self.entries["Number of Students"].get())
            
            if len(self.w_perfect) != w_count or len(self.p_perfect) != p_count or len(self.a_perfect) != a_count:
                raise ValueError("Mismatch in activity count and scores.")

        except Exception as e:
            messagebox.showerror("Input Error", str(e))
            return
        
        self.clear_window()
        
        # Main container
        main_frame = tk.Frame(self.root, bg=ModernStyle.BACKGROUND)
        main_frame.pack(expand=True, fill="both", padx=30, pady=20)
        
        # Compact header
        header_frame = tk.Frame(main_frame, bg=ModernStyle.BACKGROUND)
        header_frame.pack(fill="x", pady=(0, 15))
        
        back_btn = create_modern_button(
            header_frame, "← Back", self.build_form, 
            bg_color=ModernStyle.TEXT_SECONDARY, width=8
        )
        back_btn.pack(side="left")
        
        title_label = tk.Label(
            header_frame,
            text=f"Target Grades - {self.num_students} Students",
            font=ModernStyle.FONT_TITLE,
            bg=ModernStyle.BACKGROUND,
            fg=ModernStyle.PRIMARY
        )
        title_label.pack(side="left", padx=(15, 0))
        
        # Students card with vertical list
        card_frame = tk.Frame(main_frame, bg=ModernStyle.SURFACE, relief="solid", bd=1)
        card_frame.pack(fill="both", expand=True, pady=10)
        
        # Create left and right columns for students
        content_frame = tk.Frame(card_frame, bg=ModernStyle.SURFACE)
        content_frame.pack(padx=25, pady=20, fill="both", expand=True)
        
        # Students list header
        list_header = tk.Label(
            content_frame,
            text="🎓 Enter Target Grades for Each Student",
            font=ModernStyle.FONT_HEADER,
            bg=ModernStyle.SURFACE,
            fg=ModernStyle.ACCENT
        )
        list_header.pack(pady=(0, 15))
        
        # Create scrollable frame for many students
        students_frame = tk.Frame(content_frame, bg=ModernStyle.SURFACE)
        students_frame.pack(fill="both", expand=True)
        
        # Create columns based on number of students
        cols = 3 if self.num_students > 15 else 2 if self.num_students > 8 else 1
        
        # Create column frames
        column_frames = []
        for i in range(cols):
            col_frame = tk.Frame(students_frame, bg=ModernStyle.SURFACE)
            col_frame.pack(side="left", fill="both", expand=True, padx=(0, 10 if i < cols-1 else 0))
            column_frames.append(col_frame)
        
        # Distribute students across columns
        self.target_entries.clear()
        students_per_col = (self.num_students + cols - 1) // cols
        
        for i in range(self.num_students):
            col_index = i // students_per_col
            if col_index >= cols:
                col_index = cols - 1
                
            parent_frame = column_frames[col_index]
            
            student_frame = tk.Frame(parent_frame, bg=ModernStyle.SURFACE)
            student_frame.pack(fill="x", pady=2)
            
            # Horizontal layout for each student
            label = tk.Label(
                student_frame,
                text=f"Student {i+1}:",
                font=ModernStyle.FONT_MEDIUM,
                bg=ModernStyle.SURFACE,
                fg=ModernStyle.TEXT_PRIMARY,
                width=12,
                anchor="w"
            )
            label.pack(side="left")
            
            entry = create_modern_entry(student_frame, width=10)
            entry.pack(side="left", padx=(5, 0))
            entry.insert(0, "85.0")  # Default grade
            
            percent_label = tk.Label(
                student_frame,
                text="%",
                font=ModernStyle.FONT_MEDIUM,
                bg=ModernStyle.SURFACE,
                fg=ModernStyle.TEXT_SECONDARY
            )
            percent_label.pack(side="left", padx=(2, 0))
            
            self.target_entries.append(entry)
        
        # Generate button at bottom
        button_frame = tk.Frame(content_frame, bg=ModernStyle.SURFACE)
        button_frame.pack(pady=(20, 0))
        
        generate_btn = create_modern_button(
            button_frame, "🚀 Generate Grades", self.generate_grades, 
            bg_color=ModernStyle.SUCCESS, width=18
        )
        generate_btn.pack()
    
    def generate_grades(self):
        """Generate grades and show results"""
        try:
            results = []
            for entry in self.target_entries:
                grade = float(entry.get())
                result = find_combination(
                    self.w_perfect, self.p_perfect, self.a_perfect, 
                    self.w_weight, self.p_weight, self.a_weight, grade
                )
                if result:
                    results.append(result)
                else:
                    results.append({"Final Grade": grade, "Error": "No matching combination found."})

            self.generated_results = results
            self.show_results(results)
            
        except Exception as e:
            messagebox.showerror("Error", str(e))
    
    def show_results(self, results):
        """Show results in a compact window"""
        result_window = tk.Toplevel(self.root)
        result_window.title("Generated Grades")
        result_window.geometry("900x600")
        result_window.configure(bg=ModernStyle.BACKGROUND)
        result_window.resizable(False, False)
        
        # Center the window
        result_window.update_idletasks()
        x = (result_window.winfo_screenwidth() // 2) - (900 // 2)
        y = (result_window.winfo_screenheight() // 2) - (600 // 2)
        result_window.geometry(f"900x600+{x}+{y}")
        
        # Compact header
        header_frame = tk.Frame(result_window, bg=ModernStyle.BACKGROUND)
        header_frame.pack(fill="x", padx=25, pady=(25, 15))
        
        title_label = tk.Label(
            header_frame,
            text="Generated Grades",
            font=ModernStyle.FONT_TITLE,
            bg=ModernStyle.BACKGROUND,
            fg=ModernStyle.PRIMARY
        )
        title_label.pack(side="left")
        
        export_btn = create_modern_button(
            header_frame, "📊 Export to Excel", self.export_to_excel, 
            bg_color=ModernStyle.SUCCESS, width=15
        )
        export_btn.pack(side="right")
        
        # Results display
        results_frame = tk.Frame(result_window, bg=ModernStyle.SURFACE, relief="solid", bd=1)
        results_frame.pack(fill="both", expand=True, padx=25, pady=(0, 25))
        
        # Create text widget with scrollbar
        text_frame = tk.Frame(results_frame, bg=ModernStyle.SURFACE)
        text_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        text_widget = tk.Text(
            text_frame,
            wrap=tk.WORD,
            font=("Consolas", 9),
            bg=ModernStyle.SURFACE,
            fg=ModernStyle.TEXT_PRIMARY,
            bd=0,
            highlightthickness=0
        )
        
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Format and insert results
        msg = ""
        for i, res in enumerate(results):
            msg += f"{'='*60}\n"
            msg += f"STUDENT {i+1} - Target Grade: {res['Final Grade']}%\n"
            msg += f"{'='*60}\n"
            
            if "Error" in res:
                msg += f"❌ {res['Error']}\n\n"
            else:
                components = ["Written Works", "Performance Task", "Quarterly Assessment"]
                for comp in components:
                    scores = res[comp][0]
                    grade = res[comp][1]
                    msg += f"{comp}:\n"
                    msg += f"  Scores: {scores}\n"
                    msg += f"  Grade: {grade:.2f}%\n\n"
                msg += f"✅ Final Grade: {res['Final Grade']}%\n"
            msg += "\n"
        
        text_widget.insert("1.0", msg)
        text_widget.config(state="disabled")
    
    def export_to_excel(self):
        """Export results to Excel"""
        if not self.generated_results:
            messagebox.showwarning("Warning", "No grades generated yet!")
            return
        
        subject = self.subject_var.get()
        success = create_excel_file(
            self.generated_results, subject, self.w_perfect, self.p_perfect, self.a_perfect,
            self.w_weight, self.p_weight, self.a_weight
        )
    
    def run(self):
        """Start the application"""
        self.root.mainloop()

# ================== Launch Application ==================
if __name__ == "__main__":
    app = GradeGeneratorApp()
    app.run()